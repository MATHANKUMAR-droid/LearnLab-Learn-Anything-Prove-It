"""
datastore.py
All persistent state lives in one Excel workbook (data/learnlab.xlsx), using
openpyxl. Four sheets: Users, OTPs, Sessions, TestAttempts, Certificates.
This keeps the whole "database" human-readable and inspectable in Excel,
per the brief.
"""

import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
XLSX_PATH = os.path.join(DATA_DIR, "learnlab.xlsx")

_lock = threading.Lock()

SHEETS = {
    "Users": ["user_id", "name", "email", "password_hash", "created_at", "email_verified"],
    "OTPs": ["email", "otp_code", "purpose", "created_at", "expires_at", "used"],
    "Sessions": ["session_id", "user_id", "email", "topic", "created_at"],
    "TestAttempts": ["attempt_id", "session_id", "user_id", "email", "topic",
                      "score", "total_questions", "percentage", "passed", "completed_at"],
    "Certificates": ["certificate_id", "user_id", "email", "user_name", "topic",
                      "score_percentage", "issued_at", "file_name"],
}


def _ensure_workbook():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(XLSX_PATH):
        wb = Workbook()
        default_sheet = wb.active
        first = True
        for sheet_name, headers in SHEETS.items():
            if first:
                ws = default_sheet
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        wb.save(XLSX_PATH)


def _append_row(sheet_name: str, row: list):
    with _lock:
        _ensure_workbook()
        wb = load_workbook(XLSX_PATH)
        ws = wb[sheet_name]
        ws.append(row)
        wb.save(XLSX_PATH)


def _read_all(sheet_name: str) -> list:
    with _lock:
        _ensure_workbook()
        wb = load_workbook(XLSX_PATH)
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            rows.append(dict(zip(headers, r)))
        return rows


def _update_row(sheet_name: str, match_col: str, match_value, updates: dict):
    with _lock:
        _ensure_workbook()
        wb = load_workbook(XLSX_PATH)
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        match_idx = headers.index(match_col)
        for row in ws.iter_rows(min_row=2):
            if row[match_idx].value == match_value:
                for key, val in updates.items():
                    col_idx = headers.index(key)
                    row[col_idx].value = val
        wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# Public API used by the Flask routes
# ---------------------------------------------------------------------------

def get_user_by_email(email: str):
    for u in _read_all("Users"):
        if u["email"].lower() == email.lower():
            return u
    return None


def create_user(user_id: str, name: str, email: str, password_hash: str):
    _append_row("Users", [user_id, name, email, password_hash,
                           datetime.utcnow().isoformat(), False])


def mark_user_verified(email: str):
    _update_row("Users", "email", email, {"email_verified": True})


def save_otp(email: str, otp_code: str, purpose: str, expires_at: str):
    _append_row("OTPs", [email, otp_code, purpose, datetime.utcnow().isoformat(), expires_at, False])


def get_latest_valid_otp(email: str, purpose: str):
    rows = [r for r in _read_all("OTPs")
            if r["email"].lower() == email.lower() and r["purpose"] == purpose and not r["used"]]
    if not rows:
        return None
    return rows[-1]


def mark_otp_used(email: str, otp_code: str):
    with _lock:
        _ensure_workbook()
        wb = load_workbook(XLSX_PATH)
        ws = wb["OTPs"]
        headers = [c.value for c in ws[1]]
        email_idx, code_idx, used_idx = headers.index("email"), headers.index("otp_code"), headers.index("used")
        for row in ws.iter_rows(min_row=2):
            if row[email_idx].value.lower() == email.lower() and str(row[code_idx].value) == str(otp_code):
                row[used_idx].value = True
        wb.save(XLSX_PATH)


def create_session(session_id: str, user_id: str, email: str, topic: str):
    _append_row("Sessions", [session_id, user_id, email, topic, datetime.utcnow().isoformat()])


def record_test_attempt(attempt_id, session_id, user_id, email, topic,
                         score, total_questions, percentage, passed):
    _append_row("TestAttempts", [attempt_id, session_id, user_id, email, topic,
                                  score, total_questions, percentage, passed,
                                  datetime.utcnow().isoformat()])


def record_certificate(certificate_id, user_id, email, user_name, topic, score_percentage, file_name):
    _append_row("Certificates", [certificate_id, user_id, email, user_name, topic,
                                  score_percentage, datetime.utcnow().isoformat(), file_name])


def workbook_path() -> str:
    _ensure_workbook()
    return XLSX_PATH
